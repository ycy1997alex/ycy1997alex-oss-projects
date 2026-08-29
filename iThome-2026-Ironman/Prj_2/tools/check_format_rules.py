"""格式檢查的回歸測試。

excel_reader 是整個專案規則最密集的地方，改動它之後跑這支：

    C:\\Users\\AlexYu\\.conda\\envs\\stats\\python.exe tools\\check_format_rules.py

每個案例都在暫存目錄現做一個 xlsx，驗證該擋的有擋下來、該放行的有放行，
而且錯誤訊息要指得出問題在 Excel 的哪一格。exit code 0 表示全過。
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from regression_app.model.excel_reader import load_dataset  # noqa: E402

TMP = Path(tempfile.mkdtemp(prefix="fmt_rules_"))

DEF_ROWS = [
    ["Feature", "ID", "A", "B", "Y"],
    ["Feature_Type", "Info_01", "Input_01", "Input_02", "Result_01"],
    ["Priority", "", "High", "High", "High"],
    ["Data_Type", "None", "Continuous", "Continuous", "Continuous"],
    ["Continuity", "", "", "", ""],
    ["Category", "", "", "", ""],
    ["Note", "編號", "甲", "乙", "目標"],
]
DATA_ROWS = [["ID", "A", "B", "Y"]] + [[f"S{i:03d}", i * 1.0, i * 2.0, i * 3.0] for i in range(1, 61)]


def make(name, definition=None, data=None, extra_sheet=False) -> Path:
    path = TMP / f"{name}.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Definition"
    for row in definition if definition is not None else DEF_ROWS:
        sheet.append(row)
    if data is not None:
        data_sheet = book.create_sheet("Data")
        for row in data:
            data_sheet.append(row)
    if extra_sheet:
        book.create_sheet("我的筆記").append(["隨手寫的東西"])
    book.save(path)
    return path


def check(label: str, path: Path, expect_ok: bool, must_mention: tuple[str, ...] = ()) -> bool:
    dataset, issues = load_dataset(path)
    errors = [i.message for i in issues if i.level == "error"]
    ok = dataset is not None

    verdict = "PASS"
    if ok != expect_ok:
        verdict = f"**FAIL**  (expected ok={expect_ok}, got ok={ok})"
    missing = [kw for kw in must_mention if not any(kw in e for e in errors)]
    if missing:
        verdict = f"**FAIL**  訊息沒提到 {missing}"

    print(f"{verdict:<12} {label}")
    for message in errors:
        print(f"             ✕ {message}")
    for message in [i.message for i in issues if i.level == "warning"]:
        print(f"             ⚠ {message}")
    print()
    return verdict.startswith("PASS")


def main() -> int:
    results = []

    results.append(check("正常檔案（外加一張無關工作表）", make("ok", data=DATA_ROWS, extra_sheet=True), True))
    results.append(check("缺少 Data 工作表", make("no_data"), False, ("缺少必要的工作表",)))

    bad_type = [r[:] for r in DEF_ROWS]
    bad_type[3][2] = "Numeric"
    results.append(check("Data_Type 非法值", make("bad_type", bad_type, DATA_ROWS), False, ("Numeric",)))

    bad_role = [r[:] for r in DEF_ROWS]
    bad_role[1][2] = "Feature_01"
    results.append(check("Feature_Type 非法值", make("bad_role", bad_role, DATA_ROWS), False, ("Feature_Type",)))

    no_result = [r[:] for r in DEF_ROWS]
    no_result[1][4] = "Input_03"
    results.append(check("沒有任何 Result 欄位", make("no_result", no_result, DATA_ROWS), False, ("Result_XX",)))

    cat_result = [r[:] for r in DEF_ROWS]
    cat_result[3][4] = "Categorical"
    results.append(check("結果變項是類別型", make("cat_result", cat_result, DATA_ROWS), False, ("Categorical",)))

    dup = [r[:] for r in DEF_ROWS]
    dup[0][3] = "A"
    results.append(check("Definition 重複欄位名稱", make("dup", dup, DATA_ROWS), False, ("重複欄位名稱",)))

    short_data = [["ID", "A", "Y"]] + [[f"S{i:03d}", i * 1.0, i * 3.0] for i in range(1, 61)]
    results.append(check("Data 少了宣告過的欄位", make("missing_col", None, short_data), False, ("缺少",)))

    text_in_num = [r[:] for r in DATA_ROWS]
    text_in_num[3][2] = "不知道"
    text_in_num[7][2] = "N/A?"
    results.append(check("連續型欄位混入文字", make("text_col", None, text_in_num), False, ("無法轉成數值",)))

    few = [["ID", "A", "B", "Y"]] + [[f"S{i:03d}", i * 1.0, i * 2.0, i * 3.0] for i in range(1, 21)]
    results.append(check("樣本數過少（警告而非錯誤）", make("few_rows", None, few), True))

    multi = [r[:] for r in DEF_ROWS]
    multi[1][2] = "Feature_01"
    multi[3][3] = "Numeric"
    results.append(check("多個問題要一次列完", make("multi", multi, DATA_ROWS), False, ("Feature_Type", "Numeric")))

    print("=" * 60)
    print(f"{sum(results)} / {len(results)} 通過")
    return 0 if all(results) else 1


if __name__ == "__main__":
    sys.exit(main())
