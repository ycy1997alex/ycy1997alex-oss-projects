# -*- coding: utf-8 -*-
"""把 Scikit-learn 的 Diabetes（糖尿病病理）資料集下載下來，套進 Data_Template_Stats.xlsx 的格式。

資料來源：Efron, Hastie, Johnstone & Tibshirani (2004), "Least Angle Regression",
Annals of Statistics 32(2), 407-499。sklearn 內建的就是這份 442 x 11 的表，
原始（未標準化）檔案掛在 NCSU 的課程網站上，這支程式直接抓那一份；
抓不到就退回 sklearn 內建副本（load_diabetes(scaled=False)，數值完全相同）。

用法（MSI-Alex）：
    C:/Users/Alex/anaconda3/python.exe fetch_diabetes_to_template.py

產出（都放在本資料夾）：
    diabetes.tab.txt           原始下載檔
    Diabetes_Data_Origin.xlsx  套用樣板格式的分析檔（Definition + Data 兩張表）
    Diabetes_Data_Wrong.xlsx   同上，但 age／bmi／glu 三欄各自被打亂，列與列的對應關係是壞的
"""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
TEMPLATE = HERE.parent / "Prj_2_StatisticalAnalysis" / "Data_Template_Stats.xlsx"
RAW_URL = "https://www4.stat.ncsu.edu/~boos/var.select/diabetes.tab.txt"
RAW_FILE = HERE / "diabetes.tab.txt"
ORIGIN_FILE = HERE / "Diabetes_Data_Origin.xlsx"
WRONG_FILE = HERE / "Diabetes_Data_Wrong.xlsx"

# 故意打亂的欄位；固定亂數種子，重跑會得到同一份 Wrong 檔
SHUFFLE_COLUMNS = ["age", "bmi", "glu"]
SHUFFLE_SEED = 20260830

# 欄位名稱 -> (Feature_Type, Data_Type, Note)。順序即 Excel 的欄序。
# 欄名用 sklearn 資料集自己的名稱：s1~s6 在官方說明裡分別是 tc / ldl / hdl / tch / ltg / glu。
COLUMNS = [
    ("age", "Input_01", "Continuous", "年齡｜AGE（Age）｜基本資料，單位為歲"),
    ("sex", "Input_02", "Categorical", "性別｜SEX（Sex）｜基本資料，原始編碼 1／2，資料集未說明何者為男女"),
    ("bmi", "Input_03", "Continuous", "身體質量指數｜BMI（Body Mass Index）｜體位，單位 kg/m^2"),
    ("bp", "Input_04", "Continuous", "平均血壓｜BP（Average Blood Pressure）｜生理指標，單位 mmHg"),
    ("tc", "Input_05", "Continuous", "總膽固醇｜S1／tc（Total Serum Cholesterol）｜血液檢驗"),
    ("ldl", "Input_06", "Continuous", "低密度脂蛋白｜S2／ldl（Low-Density Lipoproteins）｜血液檢驗"),
    ("hdl", "Input_07", "Continuous", "高密度脂蛋白｜S3／hdl（High-Density Lipoproteins）｜血液檢驗，數值越高越好"),
    ("tch", "Input_08", "Continuous", "總膽固醇／HDL 比值｜S4／tch（Total Cholesterol / HDL）｜血液檢驗"),
    ("ltg", "Input_09", "Continuous", "血清三酸甘油酯對數值｜S5／ltg（log of Serum Triglycerides）｜血液檢驗"),
    ("glu", "Input_10", "Continuous", "血糖｜S6／glu（Blood Sugar Level）｜血液檢驗"),
    ("target", "Result_01", "Continuous", "一年後疾病進展量化指標｜Y（Disease Progression）｜目標變項，連續值 25–346"),
]
# 原始檔的欄名，依序對應 COLUMNS
SOURCE_COLUMNS = ["AGE", "SEX", "BMI", "BP", "S1", "S2", "S3", "S4", "S5", "S6", "Y"]


def download_raw() -> pd.DataFrame:
    """抓原始 tab 檔；抓不到就用 sklearn 內建副本重建同樣的表。"""
    try:
        resp = requests.get(RAW_URL, timeout=30)
        resp.raise_for_status()
        RAW_FILE.write_text(resp.text, encoding="utf-8", newline="")
        print(f"[download] {RAW_URL} -> {RAW_FILE.name}（{len(resp.content)} bytes）")
        return pd.read_csv(RAW_FILE, sep="\t")
    except Exception as exc:  # 沒網路、網站掛掉都走這條
        print(f"[download] 失敗（{exc}），改用 sklearn 內建副本")
        from sklearn.datasets import load_diabetes

        bunch = load_diabetes(scaled=False, as_frame=True)
        df = bunch.frame.copy()
        df.columns = [*[c.upper() for c in bunch.feature_names], "Y"]
        df.to_csv(RAW_FILE, sep="\t", index=False)
        return df[SOURCE_COLUMNS]


def build_frame(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw[SOURCE_COLUMNS].copy()
    df.columns = [name for name, *_ in COLUMNS]
    return df


def shuffle_columns(df: pd.DataFrame) -> pd.DataFrame:
    """把指定欄位各自獨立重排——每欄的分配不變，但跟同一列其他欄位的對應關係被打斷。"""
    rng = np.random.default_rng(SHUFFLE_SEED)
    wrong = df.copy()
    for name in SHUFFLE_COLUMNS:
        wrong[name] = df[name].to_numpy()[rng.permutation(len(df))]
    return wrong


def write_workbook(df: pd.DataFrame, out_file: Path) -> None:
    """以樣板檔為底改寫，這樣字型、填色、框線、凍結窗格都跟著沿用。"""
    shutil.copyfile(TEMPLATE, out_file)
    wb = load_workbook(out_file)

    _write_definition(wb["Definition"], df)
    _write_data(wb["Data"], df)

    wb.save(out_file)
    print(f"[write] {out_file.name}：Definition 7x{len(COLUMNS) + 1}、Data {len(df)}x{len(COLUMNS)}")


def _stretch(ws, last_col: int, target_col: int, rows: range) -> None:
    """把 last_col 的儲存格樣式複製到右邊新增的欄位上。"""
    for col in range(last_col + 1, target_col + 1):
        for row in rows:
            ws.cell(row=row, column=col)._style = copy(ws.cell(row=row, column=last_col)._style)


def _clear_tail(ws, first_col: int, last_col: int, rows: range) -> None:
    for col in range(first_col, last_col + 1):
        for row in rows:
            ws.cell(row=row, column=col).value = None


def _write_definition(ws, df: pd.DataFrame) -> None:
    old_last = ws.max_column
    new_last = len(COLUMNS) + 1  # A 欄是列標籤
    _stretch(ws, old_last, new_last, range(1, 8))

    for idx, (name, ftype, dtype, note) in enumerate(COLUMNS, start=2):
        series = df[name]
        continuity = f"{series.min():g}, {series.max():g}" if dtype == "Continuous" else None
        category = ", ".join(str(v) for v in sorted(series.unique())) if dtype == "Categorical" else None
        values = [name, ftype, "High", dtype, continuity, category, note]
        for row, value in enumerate(values, start=1):
            ws.cell(row=row, column=idx).value = value

    if new_last < old_last:
        _clear_tail(ws, new_last + 1, old_last, range(1, 8))


def _write_data(ws, df: pd.DataFrame) -> None:
    old_last_col, old_last_row = ws.max_column, ws.max_row
    new_last_col, new_last_row = len(COLUMNS), len(df) + 1
    _stretch(ws, old_last_col, new_last_col, range(1, max(old_last_row, new_last_row) + 1))

    for col, (name, *_rest) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col).value = name
    for row, record in enumerate(df.itertuples(index=False), start=2):
        for col, value in enumerate(record, start=1):
            ws.cell(row=row, column=col).value = value

    if new_last_row < old_last_row:
        ws.delete_rows(new_last_row + 1, old_last_row - new_last_row)
    if new_last_col < old_last_col:
        _clear_tail(ws, new_last_col + 1, old_last_col, range(1, new_last_row + 1))


def main() -> None:
    raw = download_raw()
    df = build_frame(raw)
    assert df.notna().all().all(), "原始資料不該有缺失值"
    write_workbook(df, ORIGIN_FILE)
    write_workbook(shuffle_columns(df), WRONG_FILE)


if __name__ == "__main__":
    main()
